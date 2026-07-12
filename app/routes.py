# SPDX-License-Identifier: AGPL-3.0-or-later
"""Application routes: dashboard, receipts, entities, users, settings,
search and exports."""
import csv
import io
import os
import uuid
from datetime import datetime
from functools import wraps

from flask import (Blueprint, abort, current_app, flash, jsonify, make_response,
                   redirect, render_template, request, send_from_directory,
                   url_for)
from flask_login import current_user, login_required
from werkzeug.utils import secure_filename

from . import ocr as ocr_mod
from . import sheets as sheets_mod
from .models import (BusinessEntity, Receipt, ReceiptItem, SiteSetting, User, db)
from .permissions import ROLE_TIERS, user_meets_role
from .search import search_sections

bp = Blueprint("main", __name__)

ALLOWED_IMG = {"png", "jpg", "jpeg", "gif", "webp", "heic", "bmp", "tiff", "pdf"}
ALLOWED_LOGIN_IMG = {"png", "jpg", "jpeg", "webp", "gif"}


# ── access helpers ───────────────────────────────────────────────────────
def role_required(required):
    """Decorator: gate a route behind a minimum role tier."""
    def deco(f):
        @wraps(f)
        @login_required
        def wrapper(*args, **kwargs):
            if not user_meets_role(current_user, required):
                flash("You don't have permission to do that.", "danger")
                return redirect(url_for("main.index"))
            return f(*args, **kwargs)
        return wrapper
    return deco


def _site():
    return SiteSetting.query.first()


def _f(name, default=""):
    return (request.form.get(name) or default).strip()


def _money(name):
    """Parse a money form field to a float; blank/garbage → 0.0."""
    raw = (request.form.get(name) or "").replace("$", "").replace(",", "").strip()
    try:
        return round(float(raw), 2)
    except (ValueError, TypeError):
        return 0.0


def _parse_dt(raw):
    """Parse the datetime-local input value into a datetime, or None."""
    raw = (raw or "").strip()
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _save_upload(file_storage):
    """Persist an uploaded image, returning the stored filename (or None)."""
    if not file_storage or not file_storage.filename:
        return None
    ext = file_storage.filename.rsplit(".", 1)[-1].lower() if "." in file_storage.filename else ""
    if ext not in ALLOWED_IMG:
        flash(f"Unsupported file type: .{ext}", "warning")
        return None
    stored = f"{uuid.uuid4().hex}.{ext}"
    path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored)
    file_storage.save(path)
    return stored


# ── per-user data isolation ───────────────────────────────────────────────
# Every user sees only their own receipts + businesses; admins see everything.
def _is_admin():
    return user_meets_role(current_user, "admin")


def _receipts_q():
    q = Receipt.query
    if not _is_admin():
        q = q.filter(Receipt.created_by == current_user.id)
    return q


def _entities_q():
    q = BusinessEntity.query
    if not _is_admin():
        q = q.filter(BusinessEntity.owner_id == current_user.id)
    return q


def _own_receipt_or_404(rid):
    r = db.session.get(Receipt, rid)
    if not r or (not _is_admin() and r.created_by != current_user.id):
        abort(404)
    return r


def _own_entity_or_404(eid):
    e = db.session.get(BusinessEntity, eid)
    if not e or (not _is_admin() and e.owner_id != current_user.id):
        abort(404)
    return e


def _form_entities(receipt=None):
    """Active businesses to offer on a receipt form — those of the receipt's
    owner (so an admin editing someone else's receipt still sees valid
    businesses), or the current user's when creating."""
    owner = receipt.created_by if receipt else current_user.id
    return (BusinessEntity.query.filter_by(owner_id=owner, active=True)
            .order_by(BusinessEntity.name).all())


# ── template context ─────────────────────────────────────────────────────
@bp.app_context_processor
def inject_globals():
    is_admin = user_meets_role(current_user, "admin")
    site = _site()
    ctx = {
        "site": site,
        "can_edit": user_meets_role(current_user, "editor"),
        "is_admin": is_admin,
        "sheets_configured": sheets_mod.sheets_available(),
    }
    # The Settings/Users popup is rendered into every admin page's shell, so
    # its data rides along in the shared context (only queried for admins).
    # The Settings popup (incl. the everyone-visible Account tab) renders into
    # every page's shell, so its data rides along in the shared context.
    if getattr(current_user, "is_authenticated", False):
        from . import drive as drive_mod
        ctx["drive_available"] = drive_mod.is_available()
        ctx["my_drive_connected"] = drive_mod.is_connected(current_user)
        ctx["my_drive_email"] = getattr(current_user, "drive_account_email", None)
    if is_admin:
        ctx["settings_users"] = User.query.order_by(User.username).all()
        ctx["settings_roles"] = ROLE_TIERS
        ctx["has_turnstile_secret"] = bool(site and site.turnstile_secret_key_enc)
        ctx["sheets_key_present"] = sheets_mod.sheets_available()
        ctx["login_bg_files"] = _login_bg_files()
        from . import drive as drive_mod
        try:
            ctx["drive_redirect_uri"] = drive_mod.redirect_uri()
        except Exception:
            ctx["drive_redirect_uri"] = ""
    return ctx


def _login_bg_files():
    """Sorted list of login-background image filenames on disk."""
    folder = current_app.config["LOGIN_BG_FOLDER"]
    try:
        names = [f for f in os.listdir(folder)
                 if f.rsplit(".", 1)[-1].lower() in ALLOWED_LOGIN_IMG]
    except OSError:
        return []
    return sorted(names)


@bp.before_app_request
def _touch_last_seen():
    if current_user.is_authenticated:
        # Cheap heartbeat; only write occasionally to avoid a commit per hit.
        last = getattr(current_user, "last_seen_at", None)
        if not last or (datetime.utcnow() - last).total_seconds() > 300:
            current_user.last_seen_at = datetime.utcnow()
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()


# ── dashboard ────────────────────────────────────────────────────────────
@bp.route("/")
@login_required
def index():
    receipts = _receipts_q().all()
    total_spend = sum(r.grand_total or 0 for r in receipts)
    entities = _entities_q().order_by(BusinessEntity.name).all()

    # Per-entity totals across every receipt's allocation.
    per_entity = {e.id: 0.0 for e in entities}
    unassigned = 0.0
    for r in receipts:
        alloc = r.entity_allocations()
        if not alloc:
            unassigned += r.grand_total or 0.0
        for eid, amt in alloc.items():
            per_entity[eid] = per_entity.get(eid, 0.0) + amt

    entity_totals = sorted(
        [{"entity": e, "total": round(per_entity.get(e.id, 0.0), 2)} for e in entities],
        key=lambda x: x["total"], reverse=True)

    recent = (_receipts_q()
              .order_by(Receipt.created_at.desc())
              .limit(8).all())

    return render_template(
        "dashboard.html",
        receipt_count=len(receipts),
        total_spend=round(total_spend, 2),
        total_tax=round(sum(r.tax or 0 for r in receipts), 2),
        entity_count=len(entities),
        entity_totals=entity_totals,
        unassigned=round(unassigned, 2),
        recent=recent,
    )


# ── receipts: list ───────────────────────────────────────────────────────
@bp.route("/receipts")
@login_required
def receipts():
    q = (request.args.get("q") or "").strip()
    entity_id = request.args.get("entity", type=int)

    query = _receipts_q()
    if q:
        from .search import _match, tokenize
        toks = tokenize(q)
        if toks:
            query = query.filter(_match(toks, [
                Receipt.vendor_name, Receipt.city, Receipt.state,
                Receipt.notes, Receipt.ocr_text]))
    rows = query.order_by(
        Receipt.purchased_at.desc().nullslast(),
        Receipt.created_at.desc()).all()

    if entity_id:
        rows = [r for r in rows if entity_id in r.entity_allocations()]

    return render_template(
        "receipts.html", receipts=rows, q=q,
        entities=_entities_q().order_by(BusinessEntity.name).all(),
        active_entity=entity_id)


# ── receipts: create (with optional OCR pre-fill) ────────────────────────
@bp.route("/receipts/new", methods=["GET", "POST"])
@role_required("editor")
def receipt_new():
    if request.method == "POST":
        return _save_receipt(Receipt())

    # GET: if an image was uploaded to /receipts/scan we get a prefill payload,
    # stored in a sidecar JSON file next to the image (survives multi-worker).
    prefill = {}
    ocr_text = ""
    stored = request.args.get("img")
    if stored:
        data = _read_scan_sidecar(stored)
        prefill = data.get("fields", {})
        ocr_text = data.get("text", "")
    return render_template(
        "receipt_form.html", receipt=None, prefill=prefill, ocr_text=ocr_text,
        stored_image=stored, entities=_form_entities(),
        ocr_available=ocr_mod.ocr_available())


def _scan_sidecar_path(stored):
    return os.path.join(current_app.config["UPLOAD_FOLDER"], f"{stored}.scan.json")


def _write_scan_sidecar(stored, fields, text):
    import json
    try:
        with open(_scan_sidecar_path(stored), "w", encoding="utf-8") as f:
            json.dump({"fields": fields, "text": text}, f)
    except OSError:
        pass


def _read_scan_sidecar(stored):
    """Read + delete the sidecar (one-shot hand-off from scan → form)."""
    import json
    path = _scan_sidecar_path(stored)
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, ValueError):
        data = {}
    try:
        os.remove(path)
    except OSError:
        pass
    return data


@bp.route("/receipts/scan", methods=["POST"])
@role_required("editor")
def receipt_scan():
    """Accept a photo, OCR it, stash the parsed fields, and hand off to the
    new-receipt form with everything pre-filled."""
    stored = _save_upload(request.files.get("image"))
    if not stored:
        flash("No image was uploaded.", "warning")
        return redirect(url_for("main.receipt_new"))

    path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored)
    if ocr_mod.ocr_available() and stored.rsplit(".", 1)[-1].lower() != "pdf":
        fields, text = ocr_mod.scan(path)
        _write_scan_sidecar(stored, fields, text)
        flash("Receipt scanned — review the fields below and correct anything.",
              "success")
    else:
        flash("Image saved. OCR isn't available for this file — enter the "
              "fields manually.", "warning")
    return redirect(url_for("main.receipt_new", img=stored))


@bp.route("/receipts/scan.json", methods=["POST"])
@role_required("editor")
def receipt_scan_ajax():
    """AJAX variant of the scan: save the image, OCR it, and return the parsed
    fields as JSON so the browser can populate the form without a reload."""
    stored = _save_upload(request.files.get("image"))
    if not stored:
        return jsonify(ok=False, error="No image was uploaded."), 400
    path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored)
    is_pdf = stored.rsplit(".", 1)[-1].lower() == "pdf"
    if ocr_mod.ocr_available() and not is_pdf:
        fields, text = ocr_mod.scan(path)
        return jsonify(ok=True, stored_image=stored, ocr_text=text, fields=fields,
                       image_url=url_for("main.upload", filename=stored),
                       message="Scanned — review the fields and fix anything.")
    return jsonify(ok=True, stored_image=stored, ocr_text="", fields={},
                   image_url=url_for("main.upload", filename=stored),
                   message="Image saved. OCR isn't available for this file — "
                           "enter the fields manually.")


@bp.route("/receipts/<int:rid>")
@login_required
def receipt_detail(rid):
    r = _own_receipt_or_404(rid)
    allocations = r.entity_allocations()
    ref_ids = set(allocations.keys()) | {it.entity_id for it in r.items if it.entity_id}
    entities = {e.id: e for e in
                BusinessEntity.query.filter(BusinessEntity.id.in_(ref_ids or {0})).all()}
    return render_template("receipt_detail.html", r=r, entities=entities,
                           allocations=allocations)


@bp.route("/receipts/<int:rid>/edit", methods=["GET", "POST"])
@role_required("editor")
def receipt_edit(rid):
    r = _own_receipt_or_404(rid)
    if request.method == "POST":
        return _save_receipt(r)
    prefill = {
        "vendor_name": r.vendor_name, "city": r.city, "state": r.state,
        "subtotal": r.subtotal or "", "tax": r.tax or "",
        "tax_rate": r.tax_rate if r.tax_rate not in (None, "") else "",
        "grand_total": r.grand_total or "",
        "purchased_at": r.purchased_at.strftime("%Y-%m-%dT%H:%M") if r.purchased_at else "",
        "items": [{"description": it.description, "qty": it.qty, "cost": it.cost,
                   "entity_id": it.entity_id} for it in r.items],
    }
    return render_template(
        "receipt_form.html", receipt=r, prefill=prefill, ocr_text=r.ocr_text or "",
        stored_image=r.image_filename, entities=_form_entities(r),
        ocr_available=ocr_mod.ocr_available())


def _save_receipt(r):
    """Shared create/update handler. Reads the form into ``r`` and commits."""
    is_new = r.id is None
    if is_new:
        r.created_by = current_user.id
    # Only the owner's businesses may be referenced (per-user isolation).
    owner_id = r.created_by or current_user.id
    valid_ids = {e.id for e in
                 BusinessEntity.query.filter_by(owner_id=owner_id).all()}
    r.vendor_name = _f("vendor_name")
    r.purchased_at = _parse_dt(request.form.get("purchased_at"))
    r.city = _f("city")
    r.state = _f("state")
    r.currency = _f("currency", "USD") or "USD"
    r.subtotal = _money("subtotal")
    r.tax = _money("tax")
    r.grand_total = _money("grand_total")
    rate_raw = (request.form.get("tax_rate") or "").replace("%", "").strip()
    try:
        r.tax_rate = round(float(rate_raw), 4) if rate_raw else None
    except ValueError:
        r.tax_rate = None
    r.notes = _f("notes")
    r.split_mode = _f("split_mode", "single") or "single"
    ent_id = request.form.get("entity_id", type=int)
    r.entity_id = ent_id if ent_id in valid_ids else None
    r.split_entity_ids = ",".join(
        s for s in request.form.getlist("split_entity_ids")
        if s.isdigit() and int(s) in valid_ids)

    if is_new:
        r.image_filename = _f("stored_image") or None
        r.ocr_text = request.form.get("ocr_text") or None
        db.session.add(r)
    else:
        # Allow replacing / attaching an image on edit.
        newimg = _save_upload(request.files.get("image"))
        if newimg:
            r.image_filename = newimg

    # Rebuild line items from the submitted parallel arrays.
    descs = request.form.getlist("item_desc")
    qtys = request.form.getlist("item_qty")
    costs = request.form.getlist("item_cost")
    ent_ids = request.form.getlist("item_entity")
    for it in list(r.items):
        db.session.delete(it)
    for i, desc in enumerate(descs):
        desc = desc.strip()
        cost_raw = costs[i] if i < len(costs) else ""
        if not desc and not cost_raw.strip():
            continue
        try:
            qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1.0
        except ValueError:
            qty = 1.0
        try:
            cost = float((cost_raw or "0").replace("$", "").replace(",", ""))
        except ValueError:
            cost = 0.0
        eid = None
        if i < len(ent_ids) and ent_ids[i].strip().isdigit() and int(ent_ids[i]) in valid_ids:
            eid = int(ent_ids[i])
        r.items.append(ReceiptItem(description=desc[:300], qty=qty, cost=cost,
                                   entity_id=eid))

    db.session.commit()

    # Back up the image to Google Drive (per business-entity folders) on
    # create. Best-effort: never let a Drive hiccup block the save.
    extra = ""
    if is_new:
        from . import drive as drive_mod
        if drive_mod.is_available() and drive_mod.is_connected(current_user):
            _n, dmsg = drive_mod.upload_receipt(r, current_user)
            if dmsg:
                extra = " " + dmsg
    flash("Receipt saved." + extra, "success")
    return redirect(url_for("main.receipt_detail", rid=r.id))


@bp.route("/receipts/<int:rid>/delete", methods=["POST"])
@role_required("editor")
def receipt_delete(rid):
    r = _own_receipt_or_404(rid)
    # Remove the stored image file too (best-effort).
    if r.image_filename:
        try:
            os.remove(os.path.join(current_app.config["UPLOAD_FOLDER"], r.image_filename))
        except OSError:
            pass
    db.session.delete(r)
    db.session.commit()
    flash("Receipt deleted.", "success")
    return redirect(url_for("main.receipts"))


@bp.route("/uploads/<path:filename>")
@login_required
def upload(filename):
    return send_from_directory(current_app.config["UPLOAD_FOLDER"], filename)


# ── business entities ────────────────────────────────────────────────────
@bp.route("/entities", methods=["GET"])
@login_required
def entities():
    rows = _entities_q().order_by(BusinessEntity.name).all()
    # Totals per entity for context.
    receipts = _receipts_q().all()
    totals = {e.id: 0.0 for e in rows}
    for rec in receipts:
        for eid, amt in rec.entity_allocations().items():
            totals[eid] = totals.get(eid, 0.0) + amt
    return render_template("entities.html", entities=rows,
                           totals={k: round(v, 2) for k, v in totals.items()})


@bp.route("/entities/new", methods=["POST"])
@role_required("editor")
def entity_new():
    name = _f("name")
    if not name:
        flash("Enter a business name.", "warning")
        return redirect(url_for("main.entities"))
    if (BusinessEntity.query.filter_by(owner_id=current_user.id)
            .filter(db.func.lower(BusinessEntity.name) == name.lower()).first()):
        flash("You already have a business with that name.", "warning")
        return redirect(url_for("main.entities"))
    db.session.add(BusinessEntity(name=name[:160], owner_id=current_user.id,
                                  color=_f("color", "#0b5cff") or "#0b5cff"))
    db.session.commit()
    flash("Business added.", "success")
    return redirect(url_for("main.entities"))


@bp.route("/entities/<int:eid>/edit", methods=["POST"])
@role_required("editor")
def entity_edit(eid):
    e = _own_entity_or_404(eid)
    e.name = _f("name", e.name)[:160] or e.name
    e.color = _f("color", e.color) or e.color
    e.active = request.form.get("active") == "1"
    db.session.commit()
    flash("Business updated.", "success")
    return redirect(url_for("main.entities"))


@bp.route("/entities/<int:eid>/delete", methods=["POST"])
@role_required("editor")
def entity_delete(eid):
    e = _own_entity_or_404(eid)
    db.session.delete(e)
    db.session.commit()
    flash("Business deleted. Receipts previously billed to it are now unassigned.",
          "success")
    return redirect(url_for("main.entities"))


# ── search ───────────────────────────────────────────────────────────────
@bp.route("/api/search")
@login_required
def api_search():
    raw = (request.args.get("q") or "").strip()
    if len(raw) < 2:
        return jsonify(query=raw, total=0, sections=[])
    sections = search_sections(raw, per_section=8, owner_id=_search_owner())
    total = sum(len(s["items"]) for s in sections)
    return jsonify(query=raw, total=total, sections=sections)


@bp.route("/search")
@login_required
def search_page():
    raw = (request.args.get("q") or "").strip()
    sections = (search_sections(raw, per_section=100, owner_id=_search_owner())
                if len(raw) >= 2 else [])
    total = sum(len(s["items"]) for s in sections)
    return render_template("search_results.html", q=raw, sections=sections, total=total)


def _search_owner():
    """None for admins (search all), else the current user's id."""
    return None if _is_admin() else current_user.id


# ── exports ──────────────────────────────────────────────────────────────
def _export_rows():
    """Flatten every receipt into export rows, one per entity allocation
    (so per-business totals sum correctly). Unassigned receipts get a single
    row with a blank business."""
    entities = {e.id: e for e in BusinessEntity.query.all()}
    header = ["Date", "Vendor", "City", "State", "Business", "Subtotal",
              "Tax", "Grand Total", "Allocated to Business", "Split Mode",
              "Currency", "Notes"]
    rows = []
    for r in _receipts_q().order_by(
            Receipt.purchased_at.desc().nullslast(), Receipt.created_at.desc()).all():
        when = r.purchased_at.strftime("%Y-%m-%d %H:%M") if r.purchased_at else ""
        alloc = r.entity_allocations()
        if not alloc:
            rows.append([when, r.vendor_name, r.city, r.state, "",
                         r.subtotal or 0, r.tax or 0, r.grand_total or 0, "",
                         r.split_mode, r.currency, r.notes or ""])
            continue
        for eid, amt in alloc.items():
            ent = entities.get(eid)
            rows.append([when, r.vendor_name, r.city, r.state,
                         ent.name if ent else "", r.subtotal or 0, r.tax or 0,
                         r.grand_total or 0, amt, r.split_mode, r.currency,
                         r.notes or ""])
    return header, rows


@bp.route("/export/receipts.csv")
@login_required
def export_csv():
    header, rows = _export_rows()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    w.writerows(rows)
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="dilbyrt-receipts-{datetime.utcnow():%Y%m%d}.csv"')
    resp.headers["Cache-Control"] = "no-store"
    return resp


@bp.route("/export/receipts.xlsx")
@login_required
def export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        flash("XLSX export needs the openpyxl package (not installed).", "danger")
        return redirect(url_for("main.receipts"))

    header, rows = _export_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts"
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B5CFF")
    for row in rows:
        ws.append(row)
    # Sensible column widths.
    widths = [16, 26, 16, 8, 22, 11, 9, 12, 20, 12, 9, 30]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = wdt
    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = make_response(bio.getvalue())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["Content-Disposition"] = (
        f'attachment; filename="dilbyrt-receipts-{datetime.utcnow():%Y%m%d}.xlsx"')
    resp.headers["Cache-Control"] = "no-store"
    return resp


@bp.route("/export/sheets", methods=["POST"])
@role_required("editor")
def export_sheets():
    site = _site()
    sheet_id = site.google_sheet_id if site else None
    header, rows = _export_rows()
    # openpyxl-free stringify for Sheets.
    str_rows = [[("" if c is None else c) for c in row] for row in rows]
    ok, msg = sheets_mod.push_rows(sheet_id, header, str_rows)
    flash(msg, "success" if ok else "danger")
    return redirect(url_for("main.receipts"))


# ── users (admin) ────────────────────────────────────────────────────────
@bp.route("/users")
@role_required("admin")
def users():
    # Users now live inside the Settings popup — bounce any direct hit (and the
    # post-action redirects below) to the dashboard with the modal reopened on
    # the Users tab.
    return redirect(url_for("main.index", settings="users"))


@bp.route("/users/new", methods=["POST"])
@role_required("admin")
def user_new():
    username = _f("username")
    email = _f("email")
    password = request.form.get("password") or ""
    role = _f("role", "viewer")
    if not username or not email or not password:
        flash("Username, email and password are all required.", "warning")
        return redirect(url_for("main.users"))
    if role not in {k for k, _ in ROLE_TIERS}:
        role = "viewer"
    if User.query.filter(db.func.lower(User.username) == username.lower()).first():
        flash("That username is taken.", "warning")
        return redirect(url_for("main.users"))
    u = User(username=username[:64], email=email[:255], role=role, name=_f("name"))
    u.set_password(password)
    db.session.add(u)
    db.session.commit()
    flash("User created.", "success")
    return redirect(url_for("main.users"))


@bp.route("/users/<int:uid>/edit", methods=["POST"])
@role_required("admin")
def user_edit(uid):
    u = db.session.get(User, uid) or abort(404)
    u.name = _f("name", u.name)
    u.email = _f("email", u.email)[:255] or u.email
    new_role = _f("role", u.role)
    if new_role in {k for k, _ in ROLE_TIERS}:
        u.role = new_role
    # Guard: never let the last admin demote or disable themselves out.
    admin_count = User.query.filter_by(role="admin", disabled=False).count()
    wants_disabled = request.form.get("disabled") == "1"
    if u.role != "admin" or not wants_disabled:
        u.disabled = wants_disabled and not (u.is_admin() and admin_count <= 1)
    if request.form.get("password"):
        u.set_password(request.form["password"])
    db.session.commit()
    flash("User updated.", "success")
    return redirect(url_for("main.users"))


@bp.route("/users/<int:uid>/delete", methods=["POST"])
@role_required("admin")
def user_delete(uid):
    u = db.session.get(User, uid) or abort(404)
    if u.id == current_user.id:
        flash("You can't delete your own account.", "warning")
        return redirect(url_for("main.users"))
    if u.is_admin() and User.query.filter_by(role="admin", disabled=False).count() <= 1:
        flash("You can't delete the last admin.", "warning")
        return redirect(url_for("main.users"))
    db.session.delete(u)
    db.session.commit()
    flash("User deleted.", "success")
    return redirect(url_for("main.users"))


# ── settings (admin) ─────────────────────────────────────────────────────
@bp.route("/settings", methods=["GET", "POST"])
@role_required("admin")
def settings():
    from .crypto import encrypt
    site = _site()
    if request.method == "POST":
        # Site name is fixed to "Dilbyrt" — not user-configurable.
        # Google Sheets/Drive config lives on the Google tab (see google_save).
        site.turnstile_site_key = _f("turnstile_site_key") or None
        secret = request.form.get("turnstile_secret_key") or ""
        if secret.strip():
            site.turnstile_secret_key_enc = encrypt(secret.strip())
        elif request.form.get("clear_turnstile_secret") == "1":
            site.turnstile_secret_key_enc = None

        want_enabled = request.form.get("turnstile_enabled") == "1"
        has_keys = bool(site.turnstile_site_key and site.turnstile_secret_key_enc)
        if want_enabled and not has_keys:
            flash("Turnstile needs both a site key and a secret key before it "
                  "can be enabled.", "warning")
            site.turnstile_enabled = False
        else:
            site.turnstile_enabled = want_enabled

        db.session.commit()
        flash("Settings saved.", "success")
        return redirect(url_for("main.index", settings="general"))

    # Settings is a popup now — a direct GET just opens the dashboard with the
    # modal reopened on the General tab.
    return redirect(url_for("main.index", settings="general"))


# ── Google Sheets settings (admin) ────────────────────────────────────────
@bp.route("/settings/google", methods=["POST"])
@role_required("admin")
def google_save():
    site = _site()
    site.google_sheet_id = _f("google_sheet_id") or None
    db.session.commit()
    flash("Google settings saved.", "success")
    return redirect(url_for("main.index", settings="google"))


# ── Account + per-user Google Drive connection ────────────────────────────
@bp.route("/account")
@login_required
def account():
    # Account now lives in the Settings popup — bounce to the dashboard with
    # the modal opened on the Account tab (covers stale links + OAuth returns).
    return redirect(url_for("main.index", settings="account"))


@bp.route("/account/google/connect")
@role_required("editor")
def account_google_connect():
    import secrets as _secrets
    from flask import session
    from . import drive as drive_mod
    if not drive_mod.is_available():
        flash("Google Drive backup isn't set up on this server yet.", "warning")
        return redirect(url_for("main.account"))
    state = _secrets.token_urlsafe(24)
    session["drive_oauth_state"] = state
    return redirect(drive_mod.authorize_url(state))


@bp.route("/account/google/callback")
@role_required("editor")
def account_google_callback():
    from flask import session
    from . import drive as drive_mod
    if request.args.get("error"):
        flash(f"Google connection cancelled: {request.args['error']}", "warning")
        return redirect(url_for("main.account"))
    state = request.args.get("state")
    if not state or state != session.pop("drive_oauth_state", None):
        flash("The Google connection couldn't be verified — please try again.", "danger")
        return redirect(url_for("main.account"))
    code = request.args.get("code")
    if not code:
        return redirect(url_for("main.account"))
    ok, msg = drive_mod.exchange_code(current_user, code)
    flash(msg, "success" if ok else "danger")
    return redirect(url_for("main.account"))


@bp.route("/account/google/disconnect", methods=["POST"])
@role_required("editor")
def account_google_disconnect():
    from . import drive as drive_mod
    drive_mod.disconnect(current_user)
    flash("Disconnected your Google Drive.", "success")
    return redirect(url_for("main.account"))


# ── login background images ──────────────────────────────────────────────
@bp.route("/login-bg/<path:filename>")
def login_bg(filename):
    """Serve a login-background image. Public (no auth) because the login
    screen shows it to signed-out visitors. send_from_directory guards
    against path traversal."""
    return send_from_directory(current_app.config["LOGIN_BG_FOLDER"], filename)


def _save_login_bg(fs):
    """Persist one uploaded login-background image, normalising orientation
    (EXIF) and downscaling oversized photos. Returns the stored filename or
    None. Falls back to a raw save if Pillow can't process the file."""
    if not fs or not fs.filename:
        return None
    ext = fs.filename.rsplit(".", 1)[-1].lower() if "." in fs.filename else ""
    if ext not in ALLOWED_LOGIN_IMG:
        flash(f"Skipped unsupported file: {fs.filename}", "warning")
        return None
    folder = current_app.config["LOGIN_BG_FOLDER"]
    try:
        from PIL import Image, ImageOps
        img = Image.open(fs.stream)
        img = ImageOps.exif_transpose(img)       # honour phone rotation
        img = img.convert("RGB")
        w, h = img.size
        if max(w, h) > 2000:                       # cap size for fast loads
            scale = 2000 / max(w, h)
            img = img.resize((int(w * scale), int(h * scale)))
        stored = f"{uuid.uuid4().hex}.jpg"
        img.save(os.path.join(folder, stored), "JPEG", quality=85)
        return stored
    except Exception:
        try:
            fs.stream.seek(0)
        except Exception:
            pass
        stored = f"{uuid.uuid4().hex}.{ext}"
        try:
            fs.save(os.path.join(folder, stored))
            return stored
        except Exception:
            return None


@bp.route("/settings/login-bg/upload", methods=["POST"])
@role_required("admin")
def login_bg_upload():
    saved = sum(1 for fs in request.files.getlist("images") if _save_login_bg(fs))
    if saved:
        flash(f"Added {saved} login background image"
              f"{'s' if saved != 1 else ''}.", "success")
    else:
        flash("No images were uploaded.", "warning")
    return redirect(url_for("main.index", settings="login"))


@bp.route("/settings/login-bg/delete", methods=["POST"])
@role_required("admin")
def login_bg_delete():
    name = request.form.get("filename", "")
    # Only delete a name that's actually in the folder — this validation also
    # closes off any path-traversal attempt.
    if name and name in _login_bg_files():
        try:
            os.remove(os.path.join(current_app.config["LOGIN_BG_FOLDER"], name))
            flash("Login background removed.", "success")
        except OSError:
            flash("Couldn't remove that image.", "danger")
    return redirect(url_for("main.index", settings="login"))


@bp.route("/help")
@login_required
def help_page():
    return render_template("help.html", ocr_available=ocr_mod.ocr_available())
